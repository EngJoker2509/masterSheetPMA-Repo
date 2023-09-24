import openpyxl
import pymssql
import tkinter as tk
from tkinter import messagebox


def ConnectionSQLServerAllowanceDetails():
    conn = pymssql.connect(server='T24DB-REPORTS'                           # ,user='zeppelin'
                           # ,password='zeppelin'
                           , database='His_T24_Reports')
    cursor = conn.cursor(as_dict=True)
    cursor.callproc('dbo.API_AllowanceDetails')
    rows = []
    for row in cursor:
        rows.append(list(row.values())[0:])
    # Close cursor and connection
    cursor.close()
    conn.close()
    return rows


def ConnectionSQLServerCurrentPortfolio():
    # parameter = input()
    conn = pymssql.connect(server='T24DB-REPORTS'                           # ,user='zeppelin'
                           # , password='zeppelin'
                           , database='His_T24_Reports')
    cursor = conn.cursor(as_dict=True)
    cursor.callproc('dbo.API_Get_Portfolio_Report_Data')
    rows = []
    for row in cursor:
        rows.append(row)

    # Close cursor and connection
    cursor.close()
    conn.close()
    return rows


def ConnectionSQLServerProductiveportfolioRating():
    conn = pymssql.connect(server='T24DB-REPORTS'                           # ,user='zeppelin'
                           # , password='zeppelin'
                           , database='His_T24_Reports')
    cursor = conn.cursor(as_dict=True)
    cursor.callproc('dbo.API_ProductiveportfolioRating')
    rows = []
    for row in cursor:
        rows.append(row)

    # Close cursor and connection
    cursor.close()
    conn.close()
    return rows


def ConnectionSQLServerNumberOfActiveBorrowers():
    conn = pymssql.connect(server='T24DB-REPORTS'                           # ,user='zeppelin'
                           # , password='zeppelin'
                           , database='His_T24_Reports')
    cursor = conn.cursor(as_dict=True)
    cursor.callproc('dbo.API_NumberOfActiveBorrowers')
    rows = []
    for row in cursor:
        rows.append(row)

    # Close cursor and connection
    cursor.close()
    conn.close()
    return rows


def ConnectionSQLServerReginos():
    conn = pymssql.connect(server='T24DB-REPORTS'                           # ,user='zeppelin'
                           # ,password='zeppelin'
                           , database='His_T24_Reports')
    cursor = conn.cursor(as_dict=True)
    cursor.callproc('dbo.API_Reginos')
    rows = []
    for row in cursor:
        rows.append(row)

    # Close cursor and connection
    cursor.close()
    conn.close()
    return rows


def ConnectionSQLServerHousingImprovement():
    conn = pymssql.connect(server='T24DB-REPORTS'                           # ,user='zeppelin'
                           # ,password='zeppelin'
                           , database='His_T24_Reports')
    cursor = conn.cursor(as_dict=True)
    cursor.callproc('dbo.API_HOSUING_IMPROVMENT')
    rows = []
    for row in cursor:
        rows.append(row)

    # Close cursor and connection
    cursor.close()
    conn.close()
    return rows


def ConnectionSQLServerMaturity(End_date):
    conn = pymssql.connect(server='T24DB-REPORTS'                           # ,user='zeppelin'
                           # ,password='zeppelin'
                           , database='His_T24_Reports')
    cursor = conn.cursor(as_dict=True)
    cursor.callproc('dbo.API_Maturity', (End_date,))
    rows = []
    for row in cursor:
        rows.append(list(row.values())[2:4])

    # Close cursor and connection
    cursor.close()
    conn.close()
    return rows


def ConnectionSQLServerIncomeStatment(end_Date):
    conn = pymssql.connect(server='T24DB-REPORTS'                           # ,user='zeppelin'
                           # ,password='zeppelin'
                           , database='His_T24_Reports')
    cursor = conn.cursor(as_dict=True)
    cursor.callproc('dbo.API_INCOME_STATMENT', (end_Date,))
    rows = []
    for row in cursor:
        rows.append(list(row.values())[0:])
    # Close cursor and connection
    cursor.close()
    conn.close()
    return rows


def ConnectionSQLServerbalanceSheet(END_DATE):
    conn = pymssql.connect(server='T24DB-REPORTS'                           # ,user='zeppelin'
                           # ,password='zeppelin'
                           , database='His_T24_Reports')
    cursor = conn.cursor(as_dict=True)
    cursor.callproc('dbo.API_Balance_Sheet', (END_DATE,))
    rows = []
    for row in cursor:
        rows.append(list(row.values())[0:])

    cursor.close()
    conn.close()
    return rows


def ConnectionSQLServerFacilities():
    conn = pymssql.connect(server='T24DB-REPORTS'                           # ,user='zeppelin'
                           # ,password='zeppelin'
                           , database='His_T24_Reports')
    cursor = conn.cursor(as_dict=True)
    cursor.callproc('dbo.API_Facilities',)
    rows = []
    for row in cursor:
        # print('row',row)
        rows.append(list(row.values())[0:])
    # Close cursor and connection
    cursor.close()
    conn.close()
    return rows


def ConnectionSQLServerAnnex(end_date):
    conn = pymssql.connect(server='T24DB-REPORTS'                           # ,user='zeppelin'
                           # ,password='zeppelin'
                           , database='His_T24_Reports')
    cursor = conn.cursor(as_dict=True)
    cursor.callproc('dbo.API_Annex', (end_date,))
    rows = []
    for row in cursor:
        rows.append(row)

    # print(rows)
    # Close cursor and connection
    cursor.close()
    conn.close()
    return rows


def ConnectionSQLServerAccountsReceivableAnnex(end_date):
    conn = pymssql.connect(server='T24DB-REPORTS'                           # ,user='zeppelin'
                           # ,password='zeppelin'
                           , database='His_T24_Reports')
    cursor = conn.cursor(as_dict=True)
    cursor.callproc('dbo.API_AccountsReceivable_TABLE_ANNEX', (end_date,))
    rows = []
    for row in cursor:
        rows.append(row)

    # print(rows)
    # Close cursor and connection
    cursor.close()
    conn.close()
    return rows


def ConnectionSQLServerOtherAssetsAnnex(end_date):
    conn = pymssql.connect(server='T24DB-REPORTS'                           # ,user='zeppelin'
                           # ,password='zeppelin'
                           , database='His_T24_Reports')
    cursor = conn.cursor(as_dict=True)
    cursor.callproc('dbo.API_OtherAssets_TABLE_ANNEX', (end_date,))
    rows = []
    for row in cursor:
        rows.append(row)

    # print(rows)
    # Close cursor and connection
    cursor.close()
    conn.close()
    return rows


def ConnectionSQLServerIntangibleAssetsAnnex(end_date):
    conn = pymssql.connect(server='T24DB-REPORTS'                           # ,user='zeppelin'
                           # ,password='zeppelin'
                           , database='His_T24_Reports')
    cursor = conn.cursor(as_dict=True)
    cursor.callproc('dbo.API_IntangibleAssets_TABLE_ANNEX', (end_date,))
    rows = []
    for row in cursor:
        rows.append(row)

    # print(rows)
    # Close cursor and connection
    cursor.close()
    conn.close()
    return rows


def ConnectionSQLServerFixedAssetsAnnex(end_date):
    conn = pymssql.connect(server='T24DB-REPORTS'                           # ,user='zeppelin'
                           # ,password='zeppelin'
                           , database='His_T24_Reports')
    cursor = conn.cursor(as_dict=True)
    cursor.callproc('dbo.API_FixedAssets_TABLE_ANNEX', (end_date,))
    rows = []
    for row in cursor:
        rows.append(row)

    # print(rows)
    # Close cursor and connection
    cursor.close()
    conn.close()
    return rows


def ConnectionSQLServerOtherShortTermLiabilitiesAnnex(end_date):
    conn = pymssql.connect(server='T24DB-REPORTS'                           # ,user='zeppelin'
                           # ,password='zeppelin'
                           , database='His_T24_Reports')
    cursor = conn.cursor(as_dict=True)
    cursor.callproc(
        'dbo.API_OtherShortTermLiabilities_TABLE_ANNEX', (end_date,))
    rows = []
    for row in cursor:
        rows.append(row)

    # print(rows)
    # Close cursor and connection
    cursor.close()
    conn.close()
    return rows


def ConnectionSQLServerOtherProvisionsAnnex(end_date):
    conn = pymssql.connect(server='T24DB-REPORTS'                           # ,user='zeppelin'
                           # ,password='zeppelin'
                           , database='His_T24_Reports')
    cursor = conn.cursor(as_dict=True)
    cursor.callproc(
        'dbo.API_OtherProvisions_TABLE_ANNEX', (end_date,))
    rows = []
    for row in cursor:
        rows.append(row)

    # print(rows)
    # Close cursor and connection
    cursor.close()
    conn.close()
    return rows


def ConnectionSQLServerOtherLiabilitiesAnnex(end_date):
    conn = pymssql.connect(server='T24DB-REPORTS'                           # ,user='zeppelin'
                           # ,password='zeppelin'
                           , database='His_T24_Reports')
    cursor = conn.cursor(as_dict=True)
    cursor.callproc(
        'dbo.API_OtherLiabilities_TABLE_ANNEX', (end_date,))
    rows = []
    for row in cursor:
        rows.append(row)

    # print(rows)
    # Close cursor and connection
    cursor.close()
    conn.close()
    return rows


def ConnectionSQLServerOtherReservesAnnex(end_date):
    conn = pymssql.connect(server='T24DB-REPORTS'                           # ,user='zeppelin'
                           # ,password='zeppelin'
                           , database='His_T24_Reports')
    cursor = conn.cursor(as_dict=True)
    cursor.callproc(
        'dbo.API_OtherReserves_TABLE_ANNEX', (end_date,))
    rows = []
    for row in cursor:
        rows.append(row)

    # print(rows)
    # Close cursor and connection
    cursor.close()
    conn.close()
    return rows


def ConnectionSQLServerOtherOperatingRevenuesAnnex(end_date):
    conn = pymssql.connect(server='T24DB-REPORTS'                           # ,user='zeppelin'
                           # ,password='zeppelin'
                           , database='His_T24_Reports')
    cursor = conn.cursor(as_dict=True)
    cursor.callproc(
        'dbo.API_OtherOperatingRevenues_TABLE_ANNEX', (end_date,))
    rows = []
    for row in cursor:
        rows.append(row)

    # print(rows)
    # Close cursor and connection
    cursor.close()
    conn.close()
    return rows


def ConnectionSQLServerOtherFinancialExpensesAnnex(end_date):
    conn = pymssql.connect(server='T24DB-REPORTS'                           # ,user='zeppelin'
                           # ,password='zeppelin'
                           , database='His_T24_Reports')
    cursor = conn.cursor(as_dict=True)
    cursor.callproc(
        'dbo.API_OtherFinancialExpenses_TABLE_ANNEX', (end_date,))
    rows = []
    for row in cursor:
        rows.append(row)

    # print(rows)
    # Close cursor and connection
    cursor.close()
    conn.close()
    return rows


def ConnectionSQLServerOtherAdministrativeExpensesAnnex(end_date):
    conn = pymssql.connect(server='T24DB-REPORTS'                           # ,user='zeppelin'
                           # ,password='zeppelin'
                           , database='His_T24_Reports')
    cursor = conn.cursor(as_dict=True)
    cursor.callproc(
        'dbo.API_OtherAdministrativeExpenses_TABLE_ANNEX', (end_date,))
    rows = []
    for row in cursor:
        rows.append(row)

    # print(rows)
    # Close cursor and connection
    cursor.close()
    conn.close()
    return rows

# if __name__ == "__main__":


def Main_Menu():
    try:
        End_date = str(entry.get())
        Start_date = str(entry1.get())

        print("End date :-", End_date)
        print("Start Date :-", Start_date)
        CurrentPortfolio = ConnectionSQLServerCurrentPortfolio()
        data = CurrentPortfolio

        workbook = openpyxl.load_workbook(
            'C:/Users/admin11/Desktop/PythonProgram/MasterSheet/CallReport.xlsx')

        sheet = workbook['MFIs_XX.C.Q.YY_PR']
        sheet2 = workbook['MFIs_XX.C.Q.YY_Maturity']
        sheet3 = workbook['MFIs_XX.C.Q.YY_IS']
        sheet4 = workbook['MFIs_XX.C.Q.YY_BS']
        sheet5 = workbook['MFIs_XX.C.Q.YY_Facilities']
        sheet6 = workbook['MFIs_XX.C.Q.YY_Annex']

        start_row = 10
        for i, row in enumerate(data):
            for j, (key, value) in enumerate(row.items()):
                sheet.cell(row=start_row + i, column=j + 2).value = value

        ############################################################################################

        ProductiveportfolioRating = ConnectionSQLServerProductiveportfolioRating()
        data = ProductiveportfolioRating[0:2]

        start_row = 15
        for i, row in enumerate(data):
            for j, (key, value) in enumerate(row.items()):
                sheet.cell(row=start_row + i, column=j + 2).value = value

        data = ProductiveportfolioRating[2:4]
        start_row = 20
        for i, row in enumerate(data):
            for j, (key, value) in enumerate(row.items()):
                sheet.cell(row=start_row + i, column=j + 2).value = value

        ############################################################################################

        activeBorrowers = ConnectionSQLServerNumberOfActiveBorrowers()
        data = activeBorrowers

        start_row = 25
        for i, row in enumerate(data):
            for j, (key, value) in enumerate(row.items()):
                sheet.cell(row=start_row, column=i+2).value = value

        ############################################################################################

        Reginos = ConnectionSQLServerReginos()
        data = Reginos

        start_row = 33
        for i, row in enumerate(data):
            for j, (key, value) in enumerate(row.items()):
                sheet.cell(row=start_row + i, column=j + 1).value = value

        ############################################################################################

        HousingImprovement = ConnectionSQLServerHousingImprovement()
        data = HousingImprovement
        start_row = 42
        for i, row in enumerate(data):
            for j, (key, value) in enumerate(row.items()):
                sheet.cell(row=start_row + i, column=j + 1).value = value

        ############################################################################################

        Maturity = ConnectionSQLServerMaturity(End_date)
        data = Maturity[0:8]
        start_row = 10
        for i, row in enumerate(data):
            for j in range(len(row)):
                sheet2.cell(row=start_row+i, column=j+2).value = row[j]

        data = Maturity[8:16]
        start_row = 21
        for i, row in enumerate(data):
            for j in range(len(row)):
                sheet2.cell(row=start_row+i, column=j+2).value = row[j]

        data = Maturity[16:23]
        start_row = 64
        for i, row in enumerate(data):
            for j in range(len(row)):
                sheet2.cell(row=start_row+i, column=j+2).value = row[j]

        ###########################################################################################

        AllowanceDetails = ConnectionSQLServerAllowanceDetails()
        MFIs = openpyxl.load_workbook(
            'C:/Users/admin11/Desktop/PythonProgram/MasterSheet/MFIs.xlsx')
        sheet = MFIs['MFIs_XX.PRO.Q.YY_AllowDetails']

        data = AllowanceDetails
        start_row = 9
        for i, row in enumerate(data):
            for j, value in enumerate(row):
                sheet.cell(row=start_row + i, column=j + 1).value = value
        MFIs.save(
            'C:/Users/admin11/Desktop/PythonProgram/MasterSheet/MFIs.xlsx')

        ############################################################################################

        IncomeStatment = ConnectionSQLServerIncomeStatment(end_Date=End_date)
        data = IncomeStatment

        # print(data)

        start_row = 12
        k = 0
        for i in range(len(data)+25):
            if (i in (0, 1, 3, 4, 5, 7, 8, 10, 11, 13, 15, 16, 19, 20, 22, 25, 26)):
                sheet3.cell(row=start_row + i, column=3).value = data[0][k]
                k = k+1

        ###########################################################################################

        balanceSheet = ConnectionSQLServerbalanceSheet(END_DATE=End_date)
        data = balanceSheet

        start_row = 11
        k = 0
        for i in range(0, len(data)+35, 1):
            if (i in (0, 2, 5, 6, 8, 9, 10, 11, 12, 13, 14, 16, 17, 18, 19, 20, 21, 22, 23, 24, 27, 28, 30, 31, 32, 33, 34, 35)):
                sheet4.cell(row=start_row + i, column=3).value = data[0][k]
                k = k+1

        ###########################################################################################

        Facilites = ConnectionSQLServerFacilities()
        data = Facilites

        start_row = 10

        # print(len(data))
        for i in range(0, len(data), 1):
            for j in range(0, 18, 1):
                sheet5.cell(row=start_row+i, column=j+1).value = data[i][j]

        ###########################################################################################

        AccountsReceivable = ConnectionSQLServerAccountsReceivableAnnex(
            End_date)
        start_row = 17
        for i in range(0, len(AccountsReceivable), 1):
            sheet6.cell(row=start_row+i,
                        column=1).value = AccountsReceivable[i]['Row Number']
            sheet6.cell(row=start_row+i,
                        column=2).value = AccountsReceivable[i]['DESCRIPTION']
            sheet6.cell(row=start_row+i,
                        column=3).value = AccountsReceivable[i]['TOTAL']

        ###########################################################################################

        otherAssets = ConnectionSQLServerOtherAssetsAnnex(End_date)
        start_row = 41
        for i in range(0, len(otherAssets), 1):
            sheet6.cell(row=start_row+i,
                        column=1).value = otherAssets[i]['Row Number']
            sheet6.cell(row=start_row+i,
                        column=2).value = otherAssets[i]['DESCRIPTION']
            sheet6.cell(row=start_row+i,
                        column=3).value = otherAssets[i]['TOTAL']

        ###########################################################################################
        #
        IntangibleAssets = ConnectionSQLServerIntangibleAssetsAnnex(End_date)

        sheet6.cell(row=64, column=3).value = IntangibleAssets[0]['TOTAL']
        sheet6.cell(row=70, column=3).value = IntangibleAssets[1]['TOTAL']

        ###########################################################################################

        fixedAssets = ConnectionSQLServerFixedAssetsAnnex(End_date)
        sheet6.cell(row=76, column=3).value = fixedAssets[0]['TOTAL']
        sheet6.cell(row=78, column=3).value = fixedAssets[1]['TOTAL']
        sheet6.cell(row=79, column=3).value = fixedAssets[2]['TOTAL']
        sheet6.cell(row=80, column=3).value = fixedAssets[3]['TOTAL']
        sheet6.cell(row=85, column=3).value = fixedAssets[4]['TOTAL']

        ###########################################################################################

        OtherShortTermLiabilities = ConnectionSQLServerOtherShortTermLiabilitiesAnnex(
            End_date)
        start_row = 92
        for i in range(0, len(OtherShortTermLiabilities), 1):
            sheet6.cell(row=start_row+i,
                        column=1).value = OtherShortTermLiabilities[i]['Row Number']
            sheet6.cell(row=start_row+i,
                        column=2).value = OtherShortTermLiabilities[i]['DESCRIPTION']
            sheet6.cell(row=start_row+i,
                        column=3).value = OtherShortTermLiabilities[i]['TOTAL']

        ############################################################################################

        OtherProvisions = ConnectionSQLServerOtherProvisionsAnnex(
            End_date)

        sheet6.cell(row=116,
                    column=3).value = OtherProvisions[0]['TOTAL']
        sheet6.cell(row=117,
                    column=3).value = OtherProvisions[1]['TOTAL']

        sheet6.cell(row=118,
                    column=1).value = 3
        # OtherProvisions[2]['Row Number']
        sheet6.cell(row=118,
                    column=2).value = OtherProvisions[2]['DESCRIPTION']
        sheet6.cell(row=118,
                    column=3).value = OtherProvisions[2]['TOTAL']

        sheet6.cell(row=119,
                    column=1).value = 4
        # OtherProvisions[3]['Row Number']
        sheet6.cell(row=119,
                    column=2).value = OtherProvisions[3]['DESCRIPTION']
        sheet6.cell(row=119,
                    column=3).value = OtherProvisions[3]['TOTAL']

        ############################################################################################

        OtherLiabilities = ConnectionSQLServerOtherLiabilitiesAnnex(End_date)
        start_row = 132
        for i in range(0, len(OtherLiabilities), 1):
            sheet6.cell(row=start_row+i,
                        column=1).value = OtherLiabilities[i]['Row Number']
            sheet6.cell(row=start_row+i,
                        column=2).value = OtherLiabilities[i]['DESCRIPTION']
            sheet6.cell(row=start_row+i,
                        column=3).value = OtherLiabilities[i]['TOTAL']

        ############################################################################################

        otherReserves = ConnectionSQLServerOtherReservesAnnex(End_date)
        print(otherReserves)
        if not otherReserves:
            otherReserves = [
                {
                    'Row Number': 1,
                    'DESCRIPTION': 'Geo Political Risk Reserve',
                    'TOTAL': 0.0
                },
                {

                    'Row Number': 2,
                    'DESCRIPTION': 'Counter Cyclical Risk Reserve',
                    'TOTAL': 0.0
                },
                {

                    'Row Number': 3,
                    'DESCRIPTION': 'General Risk Reserve',
                    'TOTAL': 0.0
                },
                {

                    'Row Number': 4,
                    'DESCRIPTION': 'Compulsory Reserve',
                    'TOTAL': 0.0
                },
                {

                    'Row Number': 5,
                    'DESCRIPTION': 'Reserve to Support Equity',
                    'TOTAL': 0.0
                }
            ]
            start_row = 156
            for i in range(0, len(otherReserves), 1):
                sheet6.cell(row=start_row+i,
                            column=1).value = otherReserves[i]['Row Number']
                sheet6.cell(row=start_row+i,
                            column=2).value = otherReserves[i]['DESCRIPTION']
                sheet6.cell(row=start_row+i,
                            column=3).value = otherReserves[i]['TOTAL']
        else:
            start_row = 156
            for i in range(0, len(otherReserves), 1):
                sheet6.cell(row=start_row+i,
                            column=1).value = otherReserves[i]['Row Number']
                sheet6.cell(row=start_row+i,
                            column=2).value = otherReserves[i]['DESCRIPTION']
                sheet6.cell(row=start_row+i,
                            column=3).value = otherReserves[i]['TOTAL']
        ############################################################################################

        OtherOperatingRevenues = ConnectionSQLServerOtherOperatingRevenuesAnnex(
            End_date)
        start_row = 180
        for i in range(0, len(OtherOperatingRevenues), 1):
            sheet6.cell(row=start_row+i,
                        column=1).value = OtherOperatingRevenues[i]['Row Number']
            sheet6.cell(row=start_row+i,
                        column=2).value = OtherOperatingRevenues[i]['DESCRIPTION']
            sheet6.cell(row=start_row+i,
                        column=3).value = OtherOperatingRevenues[i]['TOTAL']

        ############################################################################################

        OtherOperatingRevenues = ConnectionSQLServerOtherFinancialExpensesAnnex(
            End_date)
        start_row = 204
        for i in range(0, len(OtherOperatingRevenues), 1):
            sheet6.cell(row=start_row+i,
                        column=1).value = OtherOperatingRevenues[i]['Row Number']
            sheet6.cell(row=start_row+i,
                        column=2).value = OtherOperatingRevenues[i]['DESCRIPTION']
            sheet6.cell(row=start_row+i,
                        column=3).value = OtherOperatingRevenues[i]['TOTAL']

        ############################################################################################

        OtherAdministrativeExpenses = ConnectionSQLServerOtherAdministrativeExpensesAnnex(
            End_date)
        start_row = 228
        for i in range(0, len(OtherAdministrativeExpenses), 1):
            sheet6.cell(row=start_row+i,
                        column=1).value = OtherAdministrativeExpenses[i]['Row Number']
            sheet6.cell(row=start_row+i,
                        column=2).value = OtherAdministrativeExpenses[i]['DESCRIPTION']
            sheet6.cell(row=start_row+i,
                        column=3).value = OtherAdministrativeExpenses[i]['TOTAL']

        workbook.save(
            'C:/Users/admin11/Desktop/PythonProgram/MasterSheet/CallReport.xlsx')
        window.destroy()

        ############################################################################################
    except:
        messagebox.showerror('Error')


window = tk.Tk()
window.title("Parameter Window")

label = tk.Label(window, text="Enter End Date:")
label.pack()

entry = tk.Entry(window)
entry.pack()

label = tk.Label(window, text="Enter Start Date:")
label.pack()

entry1 = tk.Entry(window)
entry1.pack()

button = tk.Button(window, text="Execute", command=Main_Menu)
button.pack()

window.mainloop()
